import os
import io
from datetime import datetime

from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from .db import engine, Base, get_db
from .models import User, Room, MeterReading, Bill, PriceConfig
from .schemas import (
    TokenOut, UserOut, UserCreate,
    RoomIn, RoomOut,
    ReadingIn, ReadingOut,
    PriceIn, PriceOut,
    PayUpdateIn, PayBatchUpdateIn,
    BillOut,
)
from .security import hash_password, verify_password, create_access_token
from .deps import get_current_user, require_admin
from .billing import get_latest_price, generate_bill_for_room

app = FastAPI(title="房屋出租管理系统 API", version="1.0.0")

# CORS（生产建议设置为具体域名）
origins_env = os.getenv("CORS_ORIGINS", "").strip()
if origins_env:
    allow_origins = [x.strip() for x in origins_env.split(",") if x.strip()]
else:
    allow_origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

@app.on_event("startup")
def on_startup():
    # 自动建表（无数据阶段足够用；长期建议 Alembic）
    Base.metadata.create_all(bind=engine)

    # 初始化管理员
    admin_user = os.getenv("INIT_ADMIN_USERNAME", "admin")
    admin_pass = os.getenv("INIT_ADMIN_PASSWORD", "admin123456")

    # 直接使用 engine 的 Session
    from .db import SessionLocal
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.username == admin_user).first()
        if not u:
            u = User(username=admin_user, password_hash=hash_password(admin_pass), role="admin", is_active=1)
            db.add(u)
            db.commit()
        # 确保至少存在一条单价配置
        _ = get_latest_price(db)
    finally:
        db.close()

@app.get("/api/health")
def health():
    return {"ok": True}

# -------------------- Auth --------------------
@app.post("/api/auth/login", response_model=TokenOut)
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form.username).first()
    if not user or not verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if user.is_active != 1:
        raise HTTPException(status_code=403, detail="账号已禁用")
    token = create_access_token(user.username)
    return {"access_token": token}

@app.get("/api/auth/me", response_model=UserOut)
def me(user: User = Depends(get_current_user)):
    return user

# -------------------- Users (admin) --------------------
@app.get("/api/users", response_model=list[UserOut])
def list_users(_: User = Depends(require_admin), db: Session = Depends(get_db)):
    return db.query(User).order_by(User.id.desc()).all()

@app.post("/api/users", response_model=UserOut)
def create_user(payload: UserCreate, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == payload.username).first():
        raise HTTPException(400, detail="用户名已存在")
    u = User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
        is_active=1
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return u

# -------------------- Prices --------------------
@app.get("/api/prices/latest", response_model=PriceOut)
def price_latest(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    pc = get_latest_price(db)
    return {
        "id": pc.id,
        "water_price": pc.water_price,
        "elec_price": pc.elec_price,
        "gas_price": pc.gas_price,
        "property_rate": pc.property_rate,
        "effective_from": pc.effective_from.strftime("%Y-%m-%d %H:%M:%S"),
    }

@app.get("/api/prices", response_model=list[PriceOut])
def price_list(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.query(PriceConfig).order_by(PriceConfig.effective_from.desc()).all()
    return [{
        "id": x.id,
        "water_price": x.water_price,
        "elec_price": x.elec_price,
        "gas_price": x.gas_price,
        "property_rate": x.property_rate,
        "effective_from": x.effective_from.strftime("%Y-%m-%d %H:%M:%S"),
    } for x in rows]

@app.post("/api/prices", response_model=PriceOut)
def price_create(payload: PriceIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    pc = PriceConfig(**payload.model_dump())
    db.add(pc)
    db.commit()
    db.refresh(pc)
    return {
        "id": pc.id,
        "water_price": pc.water_price,
        "elec_price": pc.elec_price,
        "gas_price": pc.gas_price,
        "property_rate": pc.property_rate,
        "effective_from": pc.effective_from.strftime("%Y-%m-%d %H:%M:%S"),
    }

# -------------------- Rooms --------------------
@app.get("/api/rooms", response_model=list[RoomOut])
def list_rooms(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Room).order_by(Room.room_no.asc()).all()

@app.post("/api/rooms", response_model=RoomOut)
def create_room(payload: RoomIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if db.query(Room).filter(Room.room_no == payload.room_no).first():
        raise HTTPException(400, detail="房间号已存在")
    r = Room(**payload.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    return r

@app.put("/api/rooms/{room_id}", response_model=RoomOut)
def update_room(room_id: int, payload: RoomIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    r = db.query(Room).filter(Room.id == room_id).first()
    if not r:
        raise HTTPException(404, detail="房间不存在")
    for k, v in payload.model_dump().items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    return r

@app.delete("/api/rooms/{room_id}")
def delete_room(room_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    r = db.query(Room).filter(Room.id == room_id).first()
    if not r:
        raise HTTPException(404, detail="房间不存在")
    db.delete(r)
    db.commit()
    return {"ok": True}

# -------------------- Readings --------------------
@app.get("/api/readings", response_model=list[ReadingOut])
def list_readings(room_id: int | None = None, period: str | None = None,
                  _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(MeterReading)
    if room_id:
        q = q.filter(MeterReading.room_id == room_id)
    if period:
        q = q.filter(MeterReading.period == period)
    return q.order_by(MeterReading.period.desc()).all()

@app.post("/api/readings", response_model=ReadingOut)
def upsert_reading(payload: ReadingIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    room = db.query(Room).filter(Room.id == payload.room_id).first()
    if not room:
        raise HTTPException(404, detail="房间不存在")

    r = (
        db.query(MeterReading)
        .filter(MeterReading.room_id == payload.room_id, MeterReading.period == payload.period)
        .first()
    )
    if not r:
        r = MeterReading(**payload.model_dump())
    else:
        r.water = payload.water
        r.elec = payload.elec
        r.gas = payload.gas

    db.add(r)
    db.commit()
    db.refresh(r)
    return r

# -------------------- Bills --------------------

def _bill_to_out(db: Session, bill: Bill) -> dict:
    room_no = db.query(Room.room_no).filter(Room.id == bill.room_id).scalar() or ""
    return {
        "id": bill.id,
        "room_id": bill.room_id,
        "room_no": room_no,
        "period": bill.period,
        "rent_fee": bill.rent_fee,
        "water_used": bill.water_used,
        "elec_used": bill.elec_used,
        "gas_used": bill.gas_used,
        "water_fee": bill.water_fee,
        "elec_fee": bill.elec_fee,
        "gas_fee": bill.gas_fee,
        "property_rate": bill.property_rate,
        "property_fee": bill.property_fee,
        "total": bill.total,
        "is_paid": bill.is_paid,
        "paid_at": bill.paid_at.strftime("%Y-%m-%d %H:%M:%S") if bill.paid_at else None,
        "pay_method": bill.pay_method,
        "remark": bill.remark,
    }

@app.post("/api/bills/generate", response_model=list[BillOut])
def generate_bills(period: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rooms = db.query(Room).all()
    result = []
    for room in rooms:
        try:
            bill = generate_bill_for_room(db, room, period)
            result.append(_bill_to_out(db, bill))
        except ValueError:
            continue
    return result

@app.get("/api/bills", response_model=list[BillOut])
def list_bills(period: str | None = None, room_id: int | None = None,
              _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(Bill)
    if period:
        q = q.filter(Bill.period == period)
    if room_id:
        q = q.filter(Bill.room_id == room_id)
    rows = q.order_by(Bill.period.desc()).all()
    return [_bill_to_out(db, b) for b in rows]

@app.patch("/api/bills/{bill_id}/pay", response_model=BillOut)
def update_pay(bill_id: int, payload: PayUpdateIn,
               _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(404, detail="账单不存在")

    bill.is_paid = payload.is_paid

    if payload.is_paid == 1 and not payload.paid_at:
        raise HTTPException(400, detail="已收状态必须填写 paid_at（YYYY-MM-DD HH:MM:SS）")

    if payload.paid_at:
        try:
            bill.paid_at = datetime.strptime(payload.paid_at, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            raise HTTPException(400, detail="paid_at 格式错误，应为 YYYY-MM-DD HH:MM:SS")
    else:
        bill.paid_at = None

    bill.pay_method = payload.pay_method
    bill.remark = payload.remark

    if payload.is_paid == 0:
        bill.paid_at = None
        bill.pay_method = None
        bill.remark = None

    db.commit()
    db.refresh(bill)
    return _bill_to_out(db, bill)

@app.patch("/api/bills/pay/batch")
def batch_update_pay(payload: PayBatchUpdateIn,
                     _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.bill_ids:
        raise HTTPException(400, detail="bill_ids 不能为空")

    if payload.is_paid == 1 and not payload.paid_at:
        raise HTTPException(400, detail="已收状态必须填写 paid_at（YYYY-MM-DD HH:MM:SS）")

    paid_dt = None
    if payload.paid_at:
        try:
            paid_dt = datetime.strptime(payload.paid_at, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            raise HTTPException(400, detail="paid_at 格式错误，应为 YYYY-MM-DD HH:MM:SS")

    bills = db.query(Bill).filter(Bill.id.in_(payload.bill_ids)).all()
    if not bills:
        raise HTTPException(404, detail="未找到任何账单")

    updated = 0
    for b in bills:
        b.is_paid = payload.is_paid
        if payload.is_paid == 1:
            b.paid_at = paid_dt
            b.pay_method = payload.pay_method
            b.remark = payload.remark
        else:
            b.paid_at = None
            b.pay_method = None
            b.remark = None
        updated += 1

    db.commit()
    return {"ok": True, "updated": updated}

# -------------------- Export (period required) --------------------
@app.get("/api/bills/export/xlsx")
def export_bills_xlsx(period: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = (
        db.query(Bill, Room.room_no)
        .join(Room, Bill.room_id == Room.id)
        .filter(Bill.period == period)
        .order_by(Room.room_no.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "账单"

    headers = [
        "房间号", "月份", "房租",
        "水用量", "水费",
        "电用量", "电费",
        "气用量", "气费",
        "物业单价", "物业费",
        "合计",
        "收款状态", "收款时间", "方式", "备注"
    ]
    ws.append(headers)

    for bill, room_no in rows:
        ws.append([
            room_no, bill.period, bill.rent_fee,
            bill.water_used, bill.water_fee,
            bill.elec_used, bill.elec_fee,
            bill.gas_used, bill.gas_fee,
            bill.property_rate, bill.property_fee,
            bill.total,
            "已收" if bill.is_paid == 1 else "未收",
            bill.paid_at.strftime("%Y-%m-%d %H:%M:%S") if bill.paid_at else "",
            bill.pay_method or "",
            bill.remark or "",
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"bills_{period}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/api/bills/export/pdf")
def export_bills_pdf(period: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = (
        db.query(Bill, Room.room_no)
        .join(Room, Bill.room_id == Room.id)
        .filter(Bill.period == period)
        .order_by(Room.room_no.asc())
        .all()
    )

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 40
    c.setFont("Helvetica", 12)
    c.drawString(40, y, f"Bills Export  Period: {period}")
    y -= 25

    c.setFont("Helvetica", 9)
    for bill, room_no in rows:
        line = (
            f"{room_no} {bill.period} rent={bill.rent_fee:.2f} "
            f"w={bill.water_fee:.2f} e={bill.elec_fee:.2f} g={bill.gas_fee:.2f} "
            f"prop={bill.property_fee:.2f} total={bill.total:.2f} "
            f"{'PAID' if bill.is_paid==1 else 'UNPAID'}"
        )
        c.drawString(40, y, line)
        y -= 14
        if y < 50:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 9)

    c.save()
    buffer.seek(0)

    filename = f"bills_{period}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
